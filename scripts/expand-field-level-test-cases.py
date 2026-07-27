from copy import copy
from pathlib import Path

from openpyxl import load_workbook


SOURCE = Path(r"C:\Users\Administrator\Desktop\01-详细测试用例-更新版.xlsx")
OUTPUT = Path(r"D:\prd-baas\docs\test-cases\singapore-account-opening\01-详细测试用例-字段级.xlsx")


def numbered(items):
    return "\n".join(f"{index}. {item}" for index, item in enumerate(items, start=1))


def make_case(module, feature, control, title, precondition, data, steps, expected,
              side="客户端", case_type="字段校验", priority="P1", transition="-", note="-"):
    return {
        "端": side,
        "一级模块": module,
        "二级功能": feature,
        "关联按钮或控件": control,
        "用例标题": title,
        "用例类型": case_type,
        "优先级": priority,
        "前置条件": precondition,
        "测试数据": data,
        "操作步骤": numbered(steps),
        "预期结果": numbered(expected),
        "状态变化": transition,
        "备注": note,
    }


cases = []


def add(*args, **kwargs):
    cases.append(make_case(*args, **kwargs))


# 客户端法币转入：逐字段覆盖。
incoming_base = "用户已登录并已开通新加坡账户；进入银行电汇入金页面。"
add("法币转入", "账户字段", "选择账户", "选择账户字段展示及可选值正确", incoming_base,
    "香港账户/美国账户/新加坡账户", ["展开选择账户下拉。", "选择新加坡账户。"],
    ["下拉仅展示当前用户已开户且可入金的账户。", "新加坡账户只出现一次。", "选中后字段回显新加坡账户，收款信息和币种同步刷新。"], priority="P0")
add("法币转入", "币种字段", "选择币种", "新加坡账户选择币种字段逐项校验", incoming_base,
    "USD/CNY/SGD/AED/JPY", ["选择新加坡账户。", "展开选择币种下拉并逐项选择。"],
    ["仅展示USD、CNY、SGD、AED、JPY。", "币种代码与中文名称正确。", "选择结果与入金表单内入金币种保持一致。", "不展示HKD、EUR及禁用币种。"], priority="P0")

incoming_readonly_fields = [
    ("收款账户", "新加坡账户 USD", "账户与币种组合展示，切换币种后代码同步变化。"),
    ("收款人名称", "WAN YARA WAN", "读取用户实际配置，不展示系统模板收款人。"),
    ("银行名称", "Green Link Digital Bank Pte. Ltd.", "读取系统新加坡账户银行配置，字段只读。"),
    ("银行地址", "20 PASIR PANJANG ROAD #07-25-28 MAPLETREE BUSINESS CITY SINGAPORE 117439", "地址完整展示且不被截断，字段只读。"),
    ("网络", "SWIFT", "固定展示当前银行网络，不允许用户编辑。"),
    ("账户号码", "0950", "优先读取用户实际账户号码，不展示模板号码0454。"),
    ("BIC / SWIFT", "GLDTSGSG", "读取系统银行配置，大小写与配置一致。"),
    ("渠道提供方", "新加坡账户 VA", "账户切换后渠道提供方同步更新。"),
]
for field, value, result in incoming_readonly_fields:
    add("法币转入", "收款银行字段", field, f"新加坡账户转入-{field}字段展示正确", incoming_base,
        f"{field}={value}", ["选择新加坡账户及USD。", f"核对“{field}”字段。", "尝试点击或输入修改。"],
        [f"字段展示“{value}”。", result, "页面不提供编辑入口，提交请求时不得由客户端覆盖该字段。"], priority="P0" if field in {"收款人名称", "账户号码", "BIC / SWIFT"} else "P1")

add("法币转入", "汇款银行字段", "打款银行", "打款银行字段展示白名单银行", incoming_base,
    "HSBC Hong Kong/账户尾号8899", ["展开打款银行选择框。", "选择一条白名单银行。"],
    ["只展示当前用户已绑定且可用的汇款银行。", "银行名称和账户尾号展示正确。", "不得选择其他客户的银行账户。"], priority="P0")
add("法币转入", "汇款银行字段", "汇款账户号码", "汇款账户号码随打款银行联动", incoming_base,
    "打款银行=HSBC Hong Kong；账号尾号=8899", ["选择打款银行。", "核对汇款账户号码。", "切换另一家银行。"],
    ["账户号码自动读取所选银行，不允许手工篡改。", "切换银行后账号同步更新。", "提交记录保存银行ID及账号快照。"], priority="P0")
add("法币转入", "表单字段", "入金币种 *", "入金币种必填及账户币种联动", incoming_base,
    "SGD", ["选择新加坡账户。", "在表单选择SGD。", "观察页面顶部币种和收款信息。"],
    ["入金币种为必填。", "表单、顶部币种和收款账户均切换为SGD。", "提交记录使用SGD。"], priority="P0")
add("法币转入", "表单字段", "入金金额 *", "入金金额字段必填、范围和精度校验", incoming_base,
    "空/0/-1/100.12/超最大限额/JPY 100.5", ["依次输入测试金额。", "每次点击提交入金申请。"],
    ["空值提示请输入入金金额。", "0和负数被拦截。", "合法金额按币种精度保存。", "超限金额显示限额提示。", "JPY小数按系统币种精度规则拦截或处理，不可静默错误入账。"], priority="P0")
add("法币转入", "表单字段", "用途 *", "用途字段必填及选项校验", incoming_base,
    "Deposit/Investment/Other", ["展开用途下拉。", "逐项选择后提交。", "清空用途再次提交。"],
    ["仅展示后台启用的用途选项。", "已选用途正确写入交易记录。", "空值显示必填提示且不提交。"], priority="P1")
add("法币转入", "表单字段", "资金来源 *", "资金来源字段必填及选项校验", incoming_base,
    "Salary/Savings/Investment income", ["展开资金来源下拉。", "逐项选择。", "清空后提交。"],
    ["选项文案和值正确。", "选择结果写入申请快照。", "空值显示必填提示且不生成申请。"], priority="P0")
add("法币转入", "表单字段", "参考备注", "参考备注可选、长度及特殊字符校验", incoming_base,
    "SG deposit 20260720/中文/符号/超长文本", ["不填写备注提交一次。", "填写合法中英文和符号提交。", "输入超过字段上限的文本。"],
    ["备注为空时允许提交。", "合法内容完整保存并安全转义。", "超长内容被限制并提示，不破坏页面布局。"], priority="P2")
add("法币转入", "表单字段", "上传文件", "入金凭证上传字段校验", incoming_base,
    "JPG/PNG/PDF/超5MB/可执行文件", ["上传合法JPG或PNG。", "删除后重新上传。", "上传超限及非法格式文件。"],
    ["合法文件显示名称、大小和上传成功状态。", "支持删除及重新上传。", "超限或非法文件被拒绝并提示。", "文件不能执行脚本，下载需鉴权。"], case_type="文件/安全")
add("法币转入", "表单操作", "提交入金申请", "提交时保存所有字段快照", incoming_base,
    "新加坡账户/SGD/100/Salary/Deposit/备注", ["完整填写每个字段。", "点击提交入金申请。", "查看客户端详情和管理端记录。"],
    ["只生成一条UNDER_REVIEW记录。", "账户、币种、金额、汇款银行、用途、资金来源、备注和银行信息快照完整。", "客户端与管理端字段值一致。"], priority="P0")

incoming_detail_fields = [
    ("交易编号", "TXN-唯一编号", "编号全局唯一、可复制并可跨端检索。"),
    ("账户类型", "新加坡账户", "不得错误显示香港账户或美国账户。"),
    ("创建日期", "实际提交时间", "时间格式和时区正确。"),
    ("审核时间", "-/实际审核时间", "待处理显示-，审核完成后显示真实时间。"),
    ("手续费", "0.00 SGD", "金额和币种与申请一致。"),
    ("收款银行", "Green Link Digital Bank Pte. Ltd.", "展示交易发生时的银行快照。"),
    ("收款账户", "0950", "展示用户实际账户号码快照。"),
    ("打款银行及账号", "HSBC Hong Kong/尾号8899", "与提交时选择一致且敏感账号脱敏。"),
]
for field, value, result in incoming_detail_fields:
    add("法币转入详情", "详情字段", field, f"法币转入详情-{field}字段正确", "已提交新加坡账户SGD入金申请并进入详情。",
        f"{field}={value}", [f"定位详情中的“{field}”字段。", "与提交表单及管理端流水比对。"],
        [f"字段展示“{value}”或对应真实值。", result, "刷新页面后字段保持一致。"], priority="P0" if field in {"交易编号", "账户类型", "收款账户"} else "P1")

# 客户端法币转出：逐字段覆盖。
outgoing_base = "用户已登录并已开通新加坡账户；进入法币转出页面。"
add("法币转出", "账户字段", "选择账户", "转出选择账户字段展示及联动正确", outgoing_base,
    "新加坡账户", ["展开账户下拉。", "选择新加坡账户。"],
    ["只展示当前用户已开户且可转出的账户。", "选中后回显新加坡账户。", "币种、余额和银行地址同步刷新。"], priority="P0")
add("法币转出", "币种字段", "选择币种", "新加坡账户转出币种字段逐项校验", outgoing_base,
    "USD/CNY/SGD/AED/JPY", ["选择新加坡账户。", "展开币种并逐项选择。"],
    ["仅展示USD、CNY、SGD、AED、JPY。", "每次切换均刷新余额、金额单位和银行地址。", "不展示未支持或禁用币种。"], priority="P0")

outgoing_recipient_fields = [
    ("收款人名称", "FIDERE TRUST LIMITED", "读取所选银行地址的账户持有人。"),
    ("收款人账户", "11020160454", "账号完整或按安全规则脱敏展示。"),
    ("银行名称 / 钱包地址", "Green Link Digital Bank Pte. Ltd.", "银行名称与所选地址一致。"),
    ("国家/地区", "新加坡", "国家地区与银行地址配置一致。"),
    ("SWIFT / Routing / Network", "GLDTSGSG", "网络标识与币种和银行匹配。"),
]
for field, value, result in outgoing_recipient_fields:
    add("法币转出", "收款人字段", field, f"新加坡账户转出-{field}字段正确", outgoing_base,
        f"{field}={value}", ["选择新加坡账户及USD。", "选择对应银行地址。", f"核对“{field}”。"],
        [f"字段展示“{value}”。", result, "切换银行地址后字段同步刷新且不残留旧值。"], priority="P0" if field in {"收款人账户", "SWIFT / Routing / Network"} else "P1")

add("法币转出", "余额字段", "可用余额", "可用余额按新加坡账户和币种展示", outgoing_base,
    "SGD 74,920.50", ["选择新加坡账户。", "选择SGD。", "核对可用余额并切换到JPY。"],
    ["展示新加坡账户SGD可用余额。", "切换JPY后显示JPY余额和对应精度。", "冻结、在途或其他账户余额不得混入可用余额。"], priority="P0")
add("法币转出", "银行地址字段", "添加新银行地址", "添加新银行地址按钮进入正确流程", outgoing_base,
    "新加坡本地银行地址", ["点击添加新银行地址。", "填写并保存银行地址。", "返回转出页面。"],
    ["进入新增银行地址页面或弹窗。", "字段按国家、币种和网络校验。", "保存后仅对当前用户可见并可被选中。"], priority="P1")
add("法币转出", "银行地址字段", "搜索银行地址", "搜索银行地址字段支持关键字和空结果", outgoing_base,
    "Green/110201/NOT-EXIST", ["分别输入银行名、账号片段和不存在关键字。"],
    ["银行名和账号片段可匹配正确记录。", "搜索忽略前后空格。", "无结果显示空状态且不自动选择错误银行。"], priority="P1")
add("法币转出", "银行地址字段", "银行地址卡片", "银行地址选择状态及账户币种匹配", outgoing_base,
    "新加坡账户/USD/GLDB", ["选择新加坡账户USD。", "点击银行地址卡片。", "切换为AED。"],
    ["选中卡片有明确选中态。", "只展示支持当前账户及币种的地址。", "切换币种后重新匹配，不沿用不支持AED的地址。"], priority="P0")
add("法币转出", "表单字段", "转出金额 *", "转出金额必填、余额和精度校验", outgoing_base,
    "空/0/-1/余额+0.01/100.12/JPY 100.5", ["选择新加坡账户和对应币种。", "依次输入测试金额并提交。"],
    ["空值、0和负数均被拦截。", "金额加服务费超过可用余额时提示余额不足。", "合法金额按币种精度保存。", "JPY小数按系统规则处理。"], priority="P0")
add("法币转出", "费用字段", "服务费", "服务费字段按币种和规则计算", outgoing_base,
    "SGD金额100；服务费2", ["选择新加坡账户SGD。", "输入金额100。", "核对服务费。"],
    ["显示2.00 SGD或后台实际规则结果。", "币种单位不得固定为USD。", "修改金额或币种后服务费实时重算。"], priority="P0")
add("法币转出", "费用字段", "实际到账", "实际到账金额字段计算正确", outgoing_base,
    "转出100 SGD；服务费2 SGD", ["输入100 SGD。", "核对实际到账。", "切换币种及金额。"],
    ["实际到账显示98.00 SGD。", "计算规则为转出金额减服务费。", "结果不得为负数，切换币种后精度和单位同步。"], priority="P0")
add("法币转出", "表单字段", "用途", "转出用途字段长度和安全校验", outgoing_base,
    "Investment/中文用途/<script>/超长文本", ["分别输入合法、脚本和超长内容。", "提交并查看详情。"],
    ["合法用途完整保存。", "脚本被转义且不执行。", "超长内容按上限限制并提示。"], priority="P1")
add("法币转出", "表单字段", "备注", "转出备注字段可选及内容校验", outgoing_base,
    "空/客户指令20260720/特殊字符", ["备注为空提交。", "填写合法备注提交。"],
    ["备注为空时按业务规则允许提交。", "合法备注写入申请和审计记录。", "特殊字符安全显示。"], priority="P2")
add("法币转出", "表单操作", "提交/确认转出", "提交时保存所有转出字段快照", outgoing_base,
    "新加坡账户/AED/100/GLDB/用途/备注", ["完整填写各字段。", "点击提交并确认。", "查看客户端详情和管理端出金审批。"],
    ["只生成一条转出申请。", "账户、币种、金额、服务费、实际到账、收款银行、用途和备注完整保存。", "提交时冻结或扣减规则正确，两端字段一致。"], priority="P0")

outgoing_detail_fields = [
    ("状态", "待审核/已完成/已拒绝", "状态与管理端审核结果一致。"),
    ("转账金额", "AED 100.00", "金额及币种与申请一致。"),
    ("服务费", "AED 2.00", "展示审核时最终服务费。"),
    ("实际到账金额", "AED 98.00", "等于转账金额减最终服务费。"),
    ("账户持有人姓名", "WAN YARA WAN", "展示当前用户账户持有人快照。"),
    ("支账账户号码", "0950", "展示用户实际新加坡账户号码。"),
    ("账户类型", "新加坡账户", "不得串到香港或美国账户。"),
    ("创建日期", "实际提交时间", "格式、时区正确。"),
    ("交易编号", "唯一编号", "可复制并可在管理端检索。"),
    ("审核时间", "-/实际审核时间", "待审核显示-，终态显示真实时间。"),
    ("收款人", "FIDERE TRUST LIMITED", "与提交时银行地址快照一致。"),
    ("收款银行", "Green Link Digital Bank Pte. Ltd.", "与提交时选择一致。"),
    ("收款银行账号", "11020160454", "完整或按安全规则脱敏展示。"),
]
for field, value, result in outgoing_detail_fields:
    add("法币转出详情", "详情字段", field, f"法币转出详情-{field}字段正确", "已提交新加坡账户AED转出并进入详情。",
        f"{field}={value}", [f"定位详情中的“{field}”字段。", "与提交表单、余额和管理端记录比对。"],
        [f"字段展示“{value}”或对应真实值。", result, "刷新页面后字段保持一致。"], priority="P0" if field in {"状态", "转账金额", "实际到账金额", "交易编号", "账户类型"} else "P1")

# 管理端手动入金：逐字段覆盖。
manual_deposit_base = "运营已登录并拥有手动入金权限；已打开手动入金抽屉。"
manual_deposit_fields = [
    ("选择客户 *", "客户ID/名称/邮箱", ["字段必填并支持按客户ID、名称或邮箱检索。", "选择后只加载该客户账户。", "切换客户时清空原账户和币种。"]),
    ("选择账户 *", "新加坡账户/0950", ["仅展示所选客户已开户且可入金的账户。", "新加坡账户展示用户实际账号0950。", "未开户账户不可选择。"]),
    ("币种 *", "USD/CNY/SGD/AED/JPY", ["仅展示新加坡账户启用币种。", "切换币种后金额单位同步。", "不展示HKD和EUR。"]),
    ("打款渠道", "电汇/FPS/ACH", ["选项值与后台渠道字典一致。", "选择结果写入资金流水。", "不支持的账户币种组合应禁用。"]),
    ("入金金额 *", "空/0/-1/100.12/JPY 100.5", ["字段必填且只能输入合法金额。", "0和负数不可提交。", "按币种精度保存，金额单位不得固定为USD。"]),
    ("备注说明 *", "线下入金凭证编号SG-IN-001", ["字段必填。", "合法中英文、数字和常用符号可保存。", "超长及脚本内容被限制或安全转义。"]),
    ("上传凭证（可选）", "JPG/PNG/超10MB/非法格式", ["合法文件上传、预览和移除正常。", "不上传仍可提交。", "超限及非法格式被阻止。"]),
]
for field, value, expected in manual_deposit_fields:
    add("手动入金", "抽屉字段", field, f"手动入金-{field}字段校验", manual_deposit_base,
        f"{field}={value}", [f"操作“{field}”并输入测试数据。", "点击确认入金。", "查看错误提示或生成结果。"],
        expected + ["字段值不得串到其他客户、账户或币种。"], side="管理端", priority="P0" if "*" in field else "P1")
add("手动入金", "抽屉操作", "确认入金", "确认入金保存全部字段并更新余额", manual_deposit_base,
    "客户A/新加坡账户/SGD/100/电汇/备注", ["完整填写每个字段。", "点击确认入金。", "查询客户资产、流水和审计日志。"],
    ["只生成一条手动入金流水。", "客户、账户、币种、金额、渠道、备注和凭证快照完整。", "SGD可用余额只增加100一次。", "记录操作人和操作时间。"], side="管理端", priority="P0")
add("手动入金", "抽屉操作", "取消/X", "取消手动入金不保存任何字段", manual_deposit_base,
    "已填写全部字段但未确认", ["分别点击取消和右上角X。", "重新打开抽屉并查询流水。"],
    ["抽屉关闭。", "不生成流水且余额不变化。", "未提交字段不会错误带入下一次操作。"], side="管理端", priority="P1")

# 管理端手动出金：逐字段覆盖。
manual_withdraw_base = "运营已登录并拥有手动出金权限；已打开手动出金抽屉。"
manual_withdraw_fields = [
    ("选择客户 *", "客户ID/名称/邮箱", ["字段必填并支持检索。", "选择后仅加载该客户账户和银行账号。", "切换客户时清空原选择。"]),
    ("选择账户 *", "新加坡账户/0950", ["仅展示该客户已开户且可出金账户。", "展示用户实际账号0950。", "未开户或关闭账户不可选择。"]),
    ("币种 *", "USD/CNY/SGD/AED/JPY", ["仅展示新加坡账户启用币种。", "金额和余额单位随币种变化。", "不展示HKD和EUR。"]),
    ("打款渠道", "电汇/FPS/ACH", ["选项来自后台渠道配置。", "与账户、币种不兼容的渠道不可选。", "最终渠道写入流水。"]),
    ("银行账号 *", "客户A/GLDB/0950", ["字段必填。", "只展示当前客户可用银行账号。", "服务端拒绝其他客户银行账号ID。"]),
    ("出金金额 *", "空/0/-1/余额+0.01/100.12/JPY 100.5", ["字段必填且只能输入合法金额。", "0、负数和超过可用余额均被拦截。", "按币种精度保存，单位不得固定为USD。"]),
    ("备注说明 *", "线下出金指令SG-OUT-001", ["字段必填。", "合法内容写入流水和审计日志。", "超长及脚本内容被限制或安全转义。"]),
    ("上传凭证（可选）", "JPG/PNG/超10MB/非法格式", ["合法文件上传、预览和移除正常。", "不上传仍可提交。", "超限及非法格式被阻止。"]),
]
for field, value, expected in manual_withdraw_fields:
    add("手动出金", "抽屉字段", field, f"手动出金-{field}字段校验", manual_withdraw_base,
        f"{field}={value}", [f"操作“{field}”并输入测试数据。", "点击确认出金。", "查看错误提示或生成结果。"],
        expected + ["字段值不得串到其他客户、账户或币种。"], side="管理端", priority="P0" if "*" in field else "P1")
add("手动出金", "余额字段", "可用余额", "手动出金读取正确账户币种可用余额", manual_withdraw_base,
    "新加坡账户AED可用余额=1,000", ["选择客户、新加坡账户和AED。", "核对余额。", "分别输入1000和1000.01。"],
    ["余额来自该客户新加坡账户AED。", "输入1000按费用规则校验。", "输入1000.01提示余额不足且不生成负余额。"], side="管理端", priority="P0")
add("手动出金", "抽屉操作", "确认出金", "确认出金保存全部字段并更新余额", manual_withdraw_base,
    "客户A/新加坡账户/AED/100/电汇/银行账号0950/备注", ["完整填写每个字段。", "点击确认出金。", "查询客户资产、流水和审计日志。"],
    ["只生成一条手动出金流水。", "客户、账户、币种、金额、渠道、银行账号、备注和凭证快照完整。", "AED可用余额只扣减100一次。", "记录操作人和操作时间。"], side="管理端", priority="P0")
add("手动出金", "抽屉操作", "取消/X", "取消手动出金不保存任何字段", manual_withdraw_base,
    "已填写全部字段但未确认", ["分别点击取消和右上角X。", "重新打开抽屉并查询流水。"],
    ["抽屉关闭。", "不生成流水且余额不变化。", "未提交字段不会错误带入下一次操作。"], side="管理端", priority="P1")


workbook = load_workbook(SOURCE)
worksheet = workbook.active
headers = [cell.value for cell in worksheet[1]]
template_row = worksheet.max_row
start_sequence = int(str(worksheet.cell(template_row, 1).value).split("-")[-1])

for offset, item in enumerate(cases, start=1):
    row_index = worksheet.max_row + 1
    row_data = {"用例ID": f"SG-{start_sequence + offset:03d}", **item}
    for column_index, header in enumerate(headers, start=1):
        source_cell = worksheet.cell(template_row, column_index)
        target_cell = worksheet.cell(row_index, column_index, row_data.get(header, "-"))
        if source_cell.has_style:
            target_cell._style = copy(source_cell._style)
        target_cell.number_format = source_cell.number_format
        target_cell.font = copy(source_cell.font)
        target_cell.fill = copy(source_cell.fill)
        target_cell.border = copy(source_cell.border)
        target_cell.alignment = copy(source_cell.alignment)
        target_cell.protection = copy(source_cell.protection)
    if worksheet.row_dimensions[template_row].height is not None:
        worksheet.row_dimensions[row_index].height = worksheet.row_dimensions[template_row].height

OUTPUT.parent.mkdir(parents=True, exist_ok=True)
workbook.save(OUTPUT)
print(f"Appended {len(cases)} field-level cases; total={worksheet.max_row - 1}; output={OUTPUT}")
